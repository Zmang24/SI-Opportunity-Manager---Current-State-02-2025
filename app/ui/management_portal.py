from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
                           QTabWidget, QTableWidget, QTableWidgetItem, QComboBox,
                           QScrollArea, QFrame, QMessageBox, QLineEdit, QFormLayout,
                           QDialog, QCheckBox, QMainWindow, QHeaderView, QTextEdit, QFileDialog)
from PyQt5.QtCore import Qt, pyqtSignal
from app.database.connection import SessionLocal
from app.models.models import User, Opportunity, ActivityLog, Notification, File, FileAttachment, Attachment, Vehicle
from datetime import datetime, timedelta, timezone
import statistics
from app.ui.dashboard import DashboardWidget
from sqlalchemy import text
import openpyxl
from openpyxl.styles import Font, PatternFill
import os
import traceback

class TicketViewDialog(QDialog):
    def __init__(self, opportunity_id, current_user, parent=None):
        super().__init__(parent)
        self.opportunity_id = opportunity_id
        self.current_user = current_user
        self.initUI()
        
    def initUI(self):
        layout = QVBoxLayout()
        layout.setSpacing(20)
        
        # Get opportunity details
        db = SessionLocal()
        try:
            opportunity = db.query(Opportunity).filter(Opportunity.id == self.opportunity_id).first()
            if not opportunity:
                QMessageBox.critical(self, "Error", "Opportunity not found")
                self.reject()
                return
                
            # Title section
            title_layout = QHBoxLayout()
            
            title = QLabel(f"{opportunity.title} - {opportunity.display_title}")
            title.setStyleSheet("""
                font-size: 18px;
                font-weight: bold;
                color: white;
            """)
            title_layout.addWidget(title)
            
            # Status combo box
            status_combo = QComboBox()
            status_combo.addItems(["New", "In Progress", "Completed", "Needs Info"])
            status_combo.setCurrentText(opportunity.status)
            status_combo.setStyleSheet("""
                QComboBox {
                    background-color: #3d3d3d;
                    color: white;
                    padding: 5px 10px;
                    border: 1px solid #555;
                    border-radius: 4px;
                    min-width: 120px;
                }
            """)
            title_layout.addWidget(status_combo)
            
            layout.addLayout(title_layout)
            
            # Info grid
            info_frame = QFrame()
            info_frame.setStyleSheet("""
                QFrame {
                    background-color: #3d3d3d;
                    border-radius: 6px;
                    padding: 15px;
                }
                QLabel {
                    color: white;
                }
            """)
            info_layout = QFormLayout()
            
            # Add info fields
            creator = db.query(User).filter(User.id == opportunity.creator_id).first()
            creator_name = f"{creator.first_name} {creator.last_name}" if creator else "Unknown"
            info_layout.addRow("Created By:", QLabel(creator_name))
            
            created_at = opportunity.created_at.strftime("%Y-%m-%d %H:%M") if opportunity.created_at else "N/A"
            info_layout.addRow("Created:", QLabel(created_at))
            
            if opportunity.acceptor_id:
                acceptor = db.query(User).filter(User.id == opportunity.acceptor_id).first()
                acceptor_name = f"{acceptor.first_name} {acceptor.last_name}" if acceptor else "Unknown"
                info_layout.addRow("Assigned To:", QLabel(acceptor_name))
            
            if opportunity.response_time:
                info_layout.addRow("Response Time:", QLabel(str(opportunity.response_time)))
            
            info_frame.setLayout(info_layout)
            layout.addWidget(info_frame)
            
            # Systems section
            if opportunity.systems:
                systems_frame = QFrame()
                systems_frame.setStyleSheet("""
                    QFrame {
                        background-color: #3d3d3d;
                        border-radius: 6px;
                        padding: 15px;
                    }
                    QLabel {
                        color: white;
                    }
                """)
                systems_layout = QVBoxLayout()
                
                systems_title = QLabel("Affected Systems")
                systems_title.setStyleSheet("font-weight: bold; font-size: 14px;")
                systems_layout.addWidget(systems_title)
                
                for system in opportunity.systems:
                    system_label = QLabel(f"• {system['system']}")
                    if system.get('affected_portions'):
                        portions = ", ".join(system['affected_portions'])
                        system_label.setText(f"• {system['system']}: {portions}")
                    systems_layout.addWidget(system_label)
                
                systems_frame.setLayout(systems_layout)
                layout.addWidget(systems_frame)
            
            # Description
            if opportunity.description:
                desc_frame = QFrame()
                desc_frame.setStyleSheet("""
                    QFrame {
                        background-color: #3d3d3d;
                        border-radius: 6px;
                        padding: 15px;
                    }
                    QLabel {
                        color: white;
                    }
                """)
                desc_layout = QVBoxLayout()
                
                desc_title = QLabel("Description")
                desc_title.setStyleSheet("font-weight: bold; font-size: 14px;")
                desc_layout.addWidget(desc_title)
                
                desc_text = QLabel(opportunity.description)
                desc_text.setWordWrap(True)
                desc_layout.addWidget(desc_text)
                
                desc_frame.setLayout(desc_layout)
                layout.addWidget(desc_frame)
            
            # Close button
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(self.accept)
            close_btn.setStyleSheet("""
                QPushButton {
                    background-color: #0078d4;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #106ebe;
                }
            """)
            layout.addWidget(close_btn, alignment=Qt.AlignCenter)
            
        finally:
            db.close()
        
        self.setLayout(layout)
        self.setWindowTitle("View Ticket")
        self.resize(600, 400)
        self.setStyleSheet("background-color: #2b2b2b;")

class UserEditDialog(QDialog):
    def __init__(self, user, is_admin=False, parent=None):
        super().__init__(parent)
        self.user = user
        self.is_admin = is_admin
        self.initUI()
        
    def initUI(self):
        layout = QVBoxLayout()
        
        # Create form layout
        form = QFormLayout()
        
        # Create input fields
        self.fields = {}
        field_configs = [
            ("username", "Username:", QLineEdit),
            ("email", "Email:", QLineEdit),
            ("first_name", "First Name:", QLineEdit),
            ("last_name", "Last Name:", QLineEdit),
            ("team", "Team:", QComboBox, ["ID3", "SI", "Email", "Advanced Projects"]),
            ("role", "Role:", QComboBox, ["user", "manager", "admin"] if self.is_admin else ["user", "manager"]),
            ("is_active", "Active:", QCheckBox)
        ]
        
        for config in field_configs:
            field_id, label, widget_type, *extra = config
            if widget_type == QComboBox and extra:
                widget = widget_type()
                widget.addItems(extra[0])
                widget.setCurrentText(getattr(self.user, field_id, "user"))
            elif widget_type == QCheckBox:
                widget = widget_type()
                widget.setChecked(getattr(self.user, field_id, True))
            else:
                widget = widget_type()
                widget.setText(str(getattr(self.user, field_id, "")))
            
            self.fields[field_id] = widget
            form.addRow(label, widget)
        
        layout.addLayout(form)
        
        # Add buttons
        buttons = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        
        buttons.addWidget(save_btn)
        buttons.addWidget(cancel_btn)
        layout.addLayout(buttons)
        
        self.setLayout(layout)
        self.setWindowTitle(f"Edit User: {self.user.username}")

class ManagementPortal(QMainWindow):
    refresh_needed = pyqtSignal()
    
    def __init__(self, current_user, parent=None):
        super().__init__(parent)
        self.current_user = current_user
        self.is_admin = current_user.role == "admin"
        self.main_window = parent
        self.dashboard = None
        self.initUI()
        
    def closeEvent(self, event):
        """Handle closing of the management portal window"""
        try:
            # Clean up any resources
            if hasattr(self, 'dashboard'):
                self.dashboard.cleanup_widgets()
            
            # Close database connections
            try:
                db = SessionLocal()
                db.close()
            except:
                pass
            
            # Hide the window instead of closing it
            event.ignore()
            self.hide()
            
        except Exception as e:
            print(f"Error during cleanup: {str(e)}")
            print(traceback.format_exc())
            event.ignore()  # Still ignore the event even if there's an error
        
    def initUI(self):
        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # Set widget background color
        self.setStyleSheet("""
            QMainWindow, QWidget {
                background-color: #2b2b2b;
            }
        """)
        
        # Header
        header = QHBoxLayout()
        title = QLabel("Management Portal")
        title.setStyleSheet("""
            font-size: 24px;
            font-weight: bold;
            color: white;
        """)
        header.addWidget(title)
        
        # Create button container for refresh and export
        button_container = QHBoxLayout()
        button_container.setSpacing(8)  # Add spacing between buttons
        
        # Add refresh button
        refresh_btn = QPushButton("↻ Refresh")
        refresh_btn.clicked.connect(self.load_data)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #106ebe;
            }
        """)
        button_container.addWidget(refresh_btn)
        
        # Add Export button
        export_btn = QPushButton("Export to Excel")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        export_btn.clicked.connect(self.export_to_excel)
        button_container.addWidget(export_btn)
        
        # Add stretch to push buttons to the right
        header.addStretch()
        
        # Add button container to header
        header.addLayout(button_container)
        
        layout.addLayout(header)
        
        # Create tab widget
        tabs = QTabWidget()
        tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #3d3d3d;
                background-color: #2b2b2b;
                border-radius: 5px;
            }
            QTabBar::tab {
                background-color: #2d2d2d;
                color: #cccccc;
                padding: 8px 16px;
                border: 1px solid #3d3d3d;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                min-width: 100px;
            }
            QTabBar::tab:selected {
                background-color: #3d3d3d;
                color: white;
            }
            QTabBar::tab:hover:!selected {
                background-color: #333333;
            }
        """)
        
        # Create and store reference to dashboard
        self.dashboard = DashboardWidget(self.current_user)
        self.dashboard.refresh_needed.connect(self.load_data)  # Connect refresh signal
        
        # Team Overview Tab
        team_tab = QWidget()
        team_layout = QVBoxLayout()
        
        # Team statistics
        stats_frame = QFrame()
        stats_frame.setStyleSheet("""
            QFrame {
                background-color: #2d2d2d;
                border-radius: 8px;
                padding: 16px;
            }
        """)
        stats_layout = QHBoxLayout()
        
        # We'll add statistics cards here
        stats_cards = [
            ("Active Tickets", "🎫"),
            ("Team Members", "👥"),
            ("Avg Response Time", "⏱"),
            ("Completion Rate", "📈")
        ]
        
        for title, icon in stats_cards:
            card = QFrame()
            card.setStyleSheet("""
                QFrame {
                    background-color: #3d3d3d;
                    border-radius: 6px;
                    padding: 12px;
                    min-width: 150px;
                }
            """)
            card_layout = QVBoxLayout()
            
            icon_label = QLabel(icon)
            icon_label.setStyleSheet("font-size: 24px;")
            card_layout.addWidget(icon_label, alignment=Qt.AlignCenter)
            
            title_label = QLabel(title)
            title_label.setStyleSheet("color: #999999; font-size: 12px;")
            card_layout.addWidget(title_label, alignment=Qt.AlignCenter)
            
            value_label = QLabel("0")  # We'll update this with real data
            value_label.setStyleSheet("color: white; font-size: 20px; font-weight: bold;")
            value_label.setObjectName(f"stat_{title.lower().replace(' ', '_')}")
            card_layout.addWidget(value_label, alignment=Qt.AlignCenter)
            
            card.setLayout(card_layout)
            stats_layout.addWidget(card)
        
        stats_frame.setLayout(stats_layout)
        team_layout.addWidget(stats_frame)
        
        # Team members table
        self.team_table = QTableWidget()
        self.team_table.setStyleSheet("""
            QTableWidget {
                background-color: #2d2d2d;
                border: none;
                gridline-color: #3d3d3d;
            }
            QTableWidget::item {
                padding: 8px;
                color: white;
            }
            QHeaderView::section {
                background-color: #3d3d3d;
                color: white;
                padding: 8px;
                border: none;
            }
        """)
        self.setup_team_table()
        team_layout.addWidget(self.team_table)
        
        team_tab.setLayout(team_layout)
        tabs.addTab(team_tab, "Team Overview")
        
        # User Management Tab (Admin only)
        if self.is_admin:
            users_tab = QWidget()
            users_layout = QVBoxLayout()
            
            # Add Role Keys section
            role_keys_frame = QFrame()
            role_keys_frame.setStyleSheet("""
                QFrame {
                    background-color: #2d2d2d;
                    border-radius: 8px;
                    padding: 16px;
                    margin-bottom: 16px;
                }
                QLabel {
                    color: white;
                }
            """)
            role_keys_layout = QVBoxLayout()
            
            # Title
            role_keys_title = QLabel("Role Access Keys")
            role_keys_title.setStyleSheet("""
                font-size: 16px;
                font-weight: bold;
                margin-bottom: 8px;
            """)
            role_keys_layout.addWidget(role_keys_title)
            
            # Keys info
            keys_info = [
                ("Manager Access Key:", "MGRPROTECH9716"),
                ("Admin Access Key:", "ADMPROTECH2025")
            ]
            
            for label_text, key in keys_info:
                key_layout = QHBoxLayout()
                label = QLabel(label_text)
                label.setStyleSheet("font-weight: bold;")
                key_layout.addWidget(label)
                
                key_value = QLabel(key)
                key_value.setStyleSheet("""
                    background-color: #3d3d3d;
                    padding: 4px 8px;
                    border-radius: 4px;
                    font-family: monospace;
                """)
                key_layout.addWidget(key_value)
                
                key_layout.addStretch()
                role_keys_layout.addLayout(key_layout)
            
            # Add note
            note = QLabel("Note: These keys are used during account creation to grant elevated privileges.")
            note.setStyleSheet("color: #999999; font-style: italic; margin-top: 8px;")
            role_keys_layout.addWidget(note)
            
            role_keys_frame.setLayout(role_keys_layout)
            users_layout.addWidget(role_keys_frame)
            
            self.users_table = QTableWidget()
            self.users_table.setStyleSheet(self.team_table.styleSheet())
            self.setup_users_table()
            users_layout.addWidget(self.users_table)
            
            users_tab.setLayout(users_layout)
            tabs.addTab(users_tab, "User Management")
        
        # Opportunities Tab
        opportunities_tab = self.create_opportunities_tab()
        tabs.addTab(opportunities_tab, "Opportunities")
        
        # Add new custom vehicles tab for admin users
        if self.current_user.role == "admin":
            tabs.addTab(self.create_vehicles_tab(), "Custom Vehicles")
        
        layout.addWidget(tabs)
        
        # Set window size
        self.resize(1200, 800)
        
        # Load initial data
        self.load_data()
        
    def setup_team_table(self):
        headers = ["Name", "Role", "Active Tickets", "Completed Tickets", "Avg Response Time", "Actions"]
        self.team_table.setColumnCount(len(headers))
        self.team_table.setHorizontalHeaderLabels(headers)
        
        # Set specific column widths
        self.team_table.setColumnWidth(0, 200)  # Name
        self.team_table.setColumnWidth(1, 100)  # Role
        self.team_table.setColumnWidth(2, 100)  # Active Tickets
        self.team_table.setColumnWidth(3, 150)  # Completed Tickets - Increased width
        self.team_table.setColumnWidth(4, 150)  # Avg Response Time - Increased width
        self.team_table.setColumnWidth(5, 100)  # Actions
        
        self.team_table.horizontalHeader().setStretchLastSection(True)
        
    def setup_users_table(self):
        if not self.is_admin:
            return
            
        headers = ["Username", "Name", "Team", "Role", "Status", "Last Active", "Actions"]
        self.users_table.setColumnCount(len(headers))
        self.users_table.setHorizontalHeaderLabels(headers)
        self.users_table.horizontalHeader().setStretchLastSection(True)
        
        # Set row height
        self.users_table.verticalHeader().setDefaultSectionSize(50)  # Increase default row height
        
        # Additional styling for better spacing
        self.users_table.setStyleSheet("""
            QTableWidget {
                background-color: #2d2d2d;
                border: none;
                gridline-color: #3d3d3d;
            }
            QTableWidget::item {
                padding: 8px;
                color: white;
            }
            QHeaderView::section {
                background-color: #3d3d3d;
                color: white;
                padding: 8px;
                border: none;
            }
            QTableWidget::item:selected {
                background-color: #0078d4;
            }
        """)

    def create_opportunities_tab(self):
        """Create the opportunities management tab"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Create table
        self.opportunities_table = QTableWidget()
        self.opportunities_table.setColumnCount(9)  # Increased column count
        self.opportunities_table.setHorizontalHeaderLabels([
            "ID", "Title", "Status", "Created By", "Created Date", "Assigned To", "Completion Time", "Response Time", "Actions"
        ])
        
        # Enable row selection
        self.opportunities_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.opportunities_table.setSelectionMode(QTableWidget.SingleSelection)
        
        # Connect double-click handler
        self.opportunities_table.itemDoubleClicked.connect(self.handle_opportunity_double_click)
        
        # Set column widths
        self.opportunities_table.setColumnWidth(0, 80)   # ID
        self.opportunities_table.setColumnWidth(1, 200)  # Title
        self.opportunities_table.setColumnWidth(2, 100)  # Status
        self.opportunities_table.setColumnWidth(3, 150)  # Created By
        self.opportunities_table.setColumnWidth(4, 150)  # Created Date
        self.opportunities_table.setColumnWidth(5, 150)  # Assigned To
        self.opportunities_table.setColumnWidth(6, 150)  # Completion Time
        self.opportunities_table.setColumnWidth(7, 100)  # Response Time
        self.opportunities_table.setColumnWidth(8, 150)  # Actions

        layout.addWidget(self.opportunities_table)
        tab.setLayout(layout)
        return tab

    def handle_opportunity_double_click(self, item):
        """Handle double-click on opportunity table row"""
        row = item.row()
        opportunity_id = self.opportunities_table.item(row, 0).text()
        self.view_opportunity(opportunity_id)

    def load_opportunities(self):
        """Load opportunities into the table"""
        try:
            db = SessionLocal()
            opportunities = db.query(Opportunity).all()
            
            self.opportunities_table.setRowCount(len(opportunities))
            
            for i, opp in enumerate(opportunities):
                # ID
                id_item = QTableWidgetItem(str(opp.id))
                id_item.setFlags(id_item.flags() & ~Qt.ItemIsEditable)  # Make read-only
                self.opportunities_table.setItem(i, 0, id_item)
                
                # Title
                title_item = QTableWidgetItem(opp.title)
                title_item.setFlags(title_item.flags() & ~Qt.ItemIsEditable)
                self.opportunities_table.setItem(i, 1, title_item)
                
                # Status
                status_item = QTableWidgetItem(opp.status)
                status_item.setFlags(status_item.flags() & ~Qt.ItemIsEditable)
                self.opportunities_table.setItem(i, 2, status_item)
                
                # Created By
                creator = db.query(User).filter(User.id == opp.creator_id).first()
                creator_item = QTableWidgetItem(creator.username if creator else "Unknown")
                creator_item.setFlags(creator_item.flags() & ~Qt.ItemIsEditable)
                self.opportunities_table.setItem(i, 3, creator_item)
                
                # Created Date
                created_date = opp.created_at.strftime("%Y-%m-%d %H:%M") if opp.created_at else "N/A"
                created_date_item = QTableWidgetItem(created_date)
                created_date_item.setFlags(created_date_item.flags() & ~Qt.ItemIsEditable)
                self.opportunities_table.setItem(i, 4, created_date_item)
                
                # Assigned To
                assigned_to = "Unassigned"
                if opp.acceptor_id:
                    acceptor = db.query(User).filter(User.id == opp.acceptor_id).first()
                    if acceptor:
                        assigned_to = f"{acceptor.first_name} {acceptor.last_name}"
                assigned_item = QTableWidgetItem(assigned_to)
                assigned_item.setFlags(assigned_item.flags() & ~Qt.ItemIsEditable)
                self.opportunities_table.setItem(i, 5, assigned_item)
                
                # Completion Time
                completion_time = "N/A"
                if opp.status.lower() == "completed" and opp.completed_at:
                    completion_time = opp.completed_at.strftime("%Y-%m-%d %H:%M")
                completion_item = QTableWidgetItem(completion_time)
                completion_item.setFlags(completion_item.flags() & ~Qt.ItemIsEditable)
                self.opportunities_table.setItem(i, 6, completion_item)
                
                # Response Time
                response_time = "N/A"
                if opp.response_time:
                    days = opp.response_time.days
                    hours = opp.response_time.seconds // 3600
                    minutes = (opp.response_time.seconds % 3600) // 60
                    if days > 0:
                        response_time = f"{days}d {hours:02d}h"
                    else:
                        response_time = f"{hours:02d}h {minutes:02d}m"
                response_item = QTableWidgetItem(response_time)
                response_item.setFlags(response_item.flags() & ~Qt.ItemIsEditable)
                self.opportunities_table.setItem(i, 7, response_item)
                
                # Actions - Create widget with buttons
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(4, 0, 4, 0)
                actions_layout.setSpacing(4)
                
                # View button
                view_btn = QPushButton("View")
                view_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #0078d4;
                        color: white;
                        border: none;
                        padding: 4px 8px;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #106ebe;
                    }
                """)
                view_btn.clicked.connect(lambda checked, oid=opp.id: self.view_opportunity(oid))
                actions_layout.addWidget(view_btn)
                
                # Delete button
                delete_btn = QPushButton("Delete")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #d83b01;
                        color: white;
                        border: none;
                        padding: 4px 8px;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #ea4a1f;
                    }
                """)
                delete_btn.clicked.connect(lambda checked, oid=opp.id: self.delete_opportunity(oid))
                actions_layout.addWidget(delete_btn)
                
                self.opportunities_table.setCellWidget(i, 8, actions_widget)
                
        finally:
            db.close()

    def delete_opportunity(self, opportunity_id):
        """Delete an opportunity after confirmation"""
        reply = QMessageBox.question(
            self,
            'Confirm Deletion',
            'Are you sure you want to delete this opportunity? This action cannot be undone.',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            db = SessionLocal()
            try:
                # Get the opportunity
                opportunity = db.query(Opportunity).filter(Opportunity.id == opportunity_id).first()
                
                if not opportunity:
                    QMessageBox.warning(self, "Error", "Opportunity not found.")
                    return
                
                # Check if user has permission (admin or creator)
                if not self.is_admin and str(opportunity.creator_id) != str(self.current_user.id):
                    QMessageBox.warning(
                        self,
                        "Permission Denied",
                        "You don't have permission to delete this opportunity."
                    )
                    return

                # Delete related records in the correct order to maintain referential integrity
                
                # Delete notifications first
                db.query(Notification).filter(Notification.opportunity_id == opportunity_id).delete()
                
                # Delete activity logs
                db.query(ActivityLog).filter(ActivityLog.opportunity_id == opportunity_id).delete()
                
                # Delete files
                db.query(File).filter(File.opportunity_id == opportunity_id).delete()
                
                # Delete file attachments
                db.query(FileAttachment).filter(FileAttachment.opportunity_id == opportunity_id).delete()
                
                # Delete attachments
                db.query(Attachment).filter(Attachment.opportunity_id == opportunity_id).delete()
                
                # Create activity log for the deletion
                activity = ActivityLog(
                    user_id=self.current_user.id,
                    action="deleted",
                    details={
                        "title": opportunity.title,
                        "status": opportunity.status,
                        "deleted_at": datetime.utcnow().isoformat()
                    },
                    created_at=datetime.utcnow()
                )
                db.add(activity)
                
                # Finally delete the opportunity
                db.delete(opportunity)
                db.commit()
                
                # Refresh the table
                self.load_opportunities()
                
                QMessageBox.information(self, "Success", "Opportunity deleted successfully.")
                
            except Exception as e:
                db.rollback()
                QMessageBox.critical(self, "Error", f"Failed to delete opportunity: {str(e)}")
            finally:
                db.close()

    def load_data(self):
        """Load all data for the management portal"""
        db = SessionLocal()
        try:
            # Get team members
            if self.is_admin:
                team_members = db.query(User).all()
            else:
                team_members = db.query(User).filter(User.team == self.current_user.team).all()
            
            # Update team table
            self.team_table.setRowCount(0)
            for member in team_members:
                row = self.team_table.rowCount()
                self.team_table.insertRow(row)
                
                # Get member statistics
                active_tickets = db.query(Opportunity).filter(
                    Opportunity.acceptor_id == member.id,
                    Opportunity.status.in_(["new", "in progress"])
                ).count()
                
                completed_tickets = db.query(Opportunity).filter(
                    Opportunity.acceptor_id == member.id,
                    Opportunity.status == "completed"
                ).count()
                
                # Calculate average response time
                completed_opps = db.query(Opportunity).filter(
                    Opportunity.acceptor_id == member.id,
                    Opportunity.status == "completed",
                    Opportunity.response_time.isnot(None)
                ).all()
                
                avg_response = "N/A"
                if completed_opps:
                    response_times = [opp.response_time.total_seconds() for opp in completed_opps]
                    avg_seconds = statistics.mean(response_times)
                    avg_response = str(timedelta(seconds=int(avg_seconds)))
                
                # Add data to table
                self.team_table.setItem(row, 0, QTableWidgetItem(f"{member.first_name} {member.last_name}"))
                self.team_table.setItem(row, 1, QTableWidgetItem(member.role))
                self.team_table.setItem(row, 2, QTableWidgetItem(str(active_tickets)))
                self.team_table.setItem(row, 3, QTableWidgetItem(str(completed_tickets)))
                self.team_table.setItem(row, 4, QTableWidgetItem(avg_response))
                
                # Add action buttons
                actions_widget = QWidget()
                actions_layout = QHBoxLayout()
                actions_layout.setContentsMargins(0, 0, 0, 0)
                
                edit_btn = QPushButton("Edit")
                edit_btn.clicked.connect(lambda checked, m=member: self.edit_user(m))
                actions_layout.addWidget(edit_btn)
                
                actions_widget.setLayout(actions_layout)
                self.team_table.setCellWidget(row, 5, actions_widget)
            
            # Update users table (admin only)
            if self.is_admin:
                self.users_table.setRowCount(0)
                for user in db.query(User).all():
                    row = self.users_table.rowCount()
                    self.users_table.insertRow(row)
                    
                    self.users_table.setItem(row, 0, QTableWidgetItem(user.username))
                    self.users_table.setItem(row, 1, QTableWidgetItem(f"{user.first_name} {user.last_name}"))
                    self.users_table.setItem(row, 2, QTableWidgetItem(user.team))
                    self.users_table.setItem(row, 3, QTableWidgetItem(user.role))
                    self.users_table.setItem(row, 4, QTableWidgetItem("Active" if user.is_active else "Inactive"))
                    self.users_table.setItem(row, 5, QTableWidgetItem(
                        user.last_active.strftime("%Y-%m-%d %H:%M") if user.last_active else "Never"
                    ))
                    
                    # Add action buttons
                    actions_widget = QWidget()
                    actions_layout = QHBoxLayout()
                    actions_layout.setContentsMargins(4, 0, 4, 0)
                    actions_layout.setSpacing(4)
                    
                    # Edit button
                    edit_btn = QPushButton("Edit")
                    edit_btn.setStyleSheet("""
                        QPushButton {
                            background-color: #0078d4;
                            color: white;
                            border: none;
                            padding: 4px 8px;
                            border-radius: 4px;
                        }
                        QPushButton:hover {
                            background-color: #106ebe;
                        }
                    """)
                    edit_btn.clicked.connect(lambda checked, u=user: self.edit_user(u))
                    actions_layout.addWidget(edit_btn)
                    
                    # Delete button (don't allow deleting self or other admins)
                    if str(user.id) != str(self.current_user.id) and user.role != "admin":
                        delete_btn = QPushButton("Delete")
                        delete_btn.setStyleSheet("""
                            QPushButton {
                                background-color: #d83b01;
                                color: white;
                                border: none;
                                padding: 4px 8px;
                                border-radius: 4px;
                            }
                            QPushButton:hover {
                                background-color: #ea4a1f;
                            }
                        """)
                        delete_btn.clicked.connect(lambda checked, u=user: self.delete_user(u))
                        actions_layout.addWidget(delete_btn)
                    
                    actions_widget.setLayout(actions_layout)
                    self.users_table.setCellWidget(row, 6, actions_widget)
            
            # Update statistics
            self.update_statistics(db)
            
            # Load opportunities
            self.load_opportunities()
            
        finally:
            db.close()
            
    def update_statistics(self, db):
        """Update the statistics cards with current data"""
        try:
            # Determine filter based on admin status
            if self.is_admin:
                team_filter = True
            else:
                team_members = db.query(User).filter(User.team == self.current_user.team).all()
                team_member_ids = [m.id for m in team_members]
                team_filter = Opportunity.creator_id.in_(team_member_ids) | Opportunity.acceptor_id.in_(team_member_ids)
            
            # Active tickets (In Progress or New)
            active_tickets = db.query(Opportunity).filter(
                team_filter if self.is_admin else True,
                Opportunity.status.ilike("in progress") | Opportunity.status.ilike("new")
            ).count()
            self.findChild(QLabel, "stat_active_tickets").setText(str(active_tickets))
            
            # Team members
            team_count = db.query(User).filter(
                User.team == self.current_user.team if not self.is_admin else True,
                User.is_active == True  # Only count active team members
            ).count()
            self.findChild(QLabel, "stat_team_members").setText(str(team_count))
            
            # Average response time and completion rate
            completed_opps = db.query(Opportunity).filter(
                team_filter if self.is_admin else True,
                Opportunity.status.ilike("completed"),
                Opportunity.completed_at.isnot(None),
                Opportunity.created_at.isnot(None)
            ).all()
            
            total_tickets = db.query(Opportunity).filter(
                team_filter if self.is_admin else True
            ).count()
            
            # Calculate average response time
            if completed_opps:
                response_times = []
                for opp in completed_opps:
                    if opp.completed_at and opp.created_at:
                        duration = opp.completed_at - opp.created_at
                        response_times.append(duration.total_seconds())
                
                if response_times:
                    avg_seconds = sum(response_times) / len(response_times)
                    days = int(avg_seconds // 86400)
                    hours = int((avg_seconds % 86400) // 3600)
                    minutes = int((avg_seconds % 3600) // 60)
                    
                    if days > 0:
                        avg_response = f"{days}d {hours}h"
                    else:
                        avg_response = f"{hours}h {minutes}m"
                else:
                    avg_response = "N/A"
            else:
                avg_response = "N/A"
            
            self.findChild(QLabel, "stat_avg_response_time").setText(avg_response)
            
            # Calculate completion rate
            if total_tickets > 0:
                completion_rate = (len(completed_opps) / total_tickets) * 100
                self.findChild(QLabel, "stat_completion_rate").setText(f"{completion_rate:.1f}%")
            else:
                self.findChild(QLabel, "stat_completion_rate").setText("0.0%")
            
            # Update team table statistics
            for row in range(self.team_table.rowCount()):
                name_item = self.team_table.item(row, 0)
                if name_item:
                    name_parts = name_item.text().split()
                    if len(name_parts) >= 2:
                        user = db.query(User).filter(
                            User.first_name == name_parts[0],
                            User.last_name == name_parts[1]
                        ).first()
                        
                        if user:
                            # Count active tickets for user
                            active_count = db.query(Opportunity).filter(
                                (Opportunity.creator_id == user.id) | (Opportunity.acceptor_id == user.id),
                                Opportunity.status.in_(["new", "in progress"])
                            ).count()
                            self.team_table.setItem(row, 2, QTableWidgetItem(str(active_count)))
                            
                            # Count completed tickets
                            completed_count = db.query(Opportunity).filter(
                                (Opportunity.creator_id == user.id) | (Opportunity.acceptor_id == user.id),
                                Opportunity.status.ilike("completed")
                            ).count()
                            self.team_table.setItem(row, 3, QTableWidgetItem(str(completed_count)))
                            
                            # Calculate average response time for user
                            user_completed = db.query(Opportunity).filter(
                                (Opportunity.creator_id == user.id) | (Opportunity.acceptor_id == user.id),
                                Opportunity.status.ilike("completed"),
                                Opportunity.completed_at.isnot(None),
                                Opportunity.created_at.isnot(None)
                            ).all()
                            
                            if user_completed:
                                user_times = []
                                for opp in user_completed:
                                    if opp.completed_at and opp.created_at:
                                        duration = opp.completed_at - opp.created_at
                                        user_times.append(duration.total_seconds())
                                
                                if user_times:
                                    avg_secs = sum(user_times) / len(user_times)
                                    days = int(avg_secs // 86400)
                                    hours = int((avg_secs % 86400) // 3600)
                                    user_avg = f"{days}d {hours}h" if days > 0 else f"{hours}h"
                                    self.team_table.setItem(row, 4, QTableWidgetItem(user_avg))
                                else:
                                    self.team_table.setItem(row, 4, QTableWidgetItem("N/A"))
                            else:
                                self.team_table.setItem(row, 4, QTableWidgetItem("N/A"))
            
        except Exception as e:
            print(f"Error updating statistics: {str(e)}")
            print(traceback.format_exc())
            
    def edit_user(self, user):
        """Open dialog to edit user details"""
        dialog = UserEditDialog(user, self.is_admin, self)
        if dialog.exec_():
            db = SessionLocal()
            try:
                # Update user with new values
                user = db.query(User).filter(User.id == user.id).first()
                for field_id, widget in dialog.fields.items():
                    if isinstance(widget, QCheckBox):
                        value = widget.isChecked()
                    else:
                        value = widget.currentText() if isinstance(widget, QComboBox) else widget.text()
                    setattr(user, field_id, value)
                
                user.updated_at = datetime.utcnow()
                db.commit()
                
                # Refresh the display
                self.load_data()
                self.refresh_needed.emit()
                
            except Exception as e:
                db.rollback()
                QMessageBox.critical(self, "Error", f"Failed to update user: {str(e)}")
            finally:
                db.close()
                
    def view_opportunity(self, opportunity_id):
        """View a specific opportunity in a dialog"""
        dialog = TicketViewDialog(opportunity_id, self.current_user, self)
        dialog.exec_()

    def delete_user(self, user):
        """Delete a user after confirmation"""
        # Don't allow deleting self or other admins
        if str(user.id) == str(self.current_user.id):
            QMessageBox.warning(self, "Error", "You cannot delete your own account.")
            return
        
        if user.role == "admin":
            QMessageBox.warning(self, "Error", "Admin accounts cannot be deleted.")
            return
        
        # Check if user has active tickets
        db = SessionLocal()
        try:
            active_tickets = db.query(Opportunity).filter(
                (Opportunity.creator_id == user.id) | (Opportunity.acceptor_id == user.id),
                Opportunity.status.in_(["new", "in progress"])
            ).count()
            
            if active_tickets > 0:
                reply = QMessageBox.question(
                    self,
                    'Warning',
                    f'This user has {active_tickets} active tickets. Are you sure you want to delete this user?',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
            else:
                reply = QMessageBox.question(
                    self,
                    'Confirm Deletion',
                    f'Are you sure you want to delete user {user.first_name} {user.last_name}?',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
            
            if reply == QMessageBox.Yes:
                # Create activity log for the deletion
                activity = ActivityLog(
                    user_id=self.current_user.id,
                    action="user_deleted",
                    details={
                        "deleted_user": {
                            "id": str(user.id),
                            "username": user.username,
                            "name": f"{user.first_name} {user.last_name}",
                            "role": user.role,
                            "team": user.team
                        },
                        "deleted_at": datetime.utcnow().isoformat()
                    },
                    created_at=datetime.utcnow()
                )
                db.add(activity)
                
                # Delete user's notifications
                db.query(Notification).filter(Notification.user_id == user.id).delete()
                
                # Delete user's activity logs
                db.query(ActivityLog).filter(ActivityLog.user_id == user.id).delete()
                
                # Delete the user
                db.delete(user)
                db.commit()
                
                QMessageBox.information(self, "Success", "User deleted successfully.")
                self.load_data()  # Refresh the display
                
        except Exception as e:
            db.rollback()
            QMessageBox.critical(self, "Error", f"Failed to delete user: {str(e)}")
        finally:
            db.close() 

    def create_vehicles_tab(self):
        """Create the custom vehicles management tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(20)
        
        # Header section
        header = QHBoxLayout()
        title = QLabel("Custom Vehicle Management")
        title.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: white;
        """)
        header.addWidget(title)
        
        refresh_btn = QPushButton("↻ Refresh")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #2d2d2d;
                color: #0078d4;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #333333;
                color: #2196F3;
            }
        """)
        refresh_btn.clicked.connect(self.load_custom_vehicles)
        header.addWidget(refresh_btn, alignment=Qt.AlignRight)
        layout.addLayout(header)
        
        # Create table
        self.vehicles_table = QTableWidget()
        self.vehicles_table.setStyleSheet("""
            QTableWidget {
                background-color: #2b2b2b;
                color: white;
                gridline-color: #3d3d3d;
                border: none;
            }
            QTableWidget::item {
                padding: 12px 8px;
                border-bottom: 1px solid #3d3d3d;
            }
            QHeaderView::section {
                background-color: #2d2d2d;
                color: white;
                padding: 12px 8px;
                border: none;
                border-right: 1px solid #3d3d3d;
                border-bottom: 1px solid #3d3d3d;
                font-weight: bold;
            }
            QTableWidget::item:selected {
                background-color: #0078d4;
            }
        """)
        
        # Set up columns
        columns = ["Year", "Make", "Model", "Created By", "Created At", "Last Modified", "Notes", "Actions"]
        self.vehicles_table.setColumnCount(len(columns))
        self.vehicles_table.setHorizontalHeaderLabels(columns)
        
        # Set column widths
        self.vehicles_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.vehicles_table.horizontalHeader().setStretchLastSection(True)
        self.vehicles_table.setColumnWidth(0, 80)  # Year
        self.vehicles_table.setColumnWidth(1, 120)  # Make
        self.vehicles_table.setColumnWidth(2, 120)  # Model
        self.vehicles_table.setColumnWidth(3, 150)  # Created By
        self.vehicles_table.setColumnWidth(4, 150)  # Created At
        self.vehicles_table.setColumnWidth(5, 150)  # Last Modified
        self.vehicles_table.setColumnWidth(6, 200)  # Notes
        
        # Set row height
        self.vehicles_table.verticalHeader().setDefaultSectionSize(50)
        self.vehicles_table.verticalHeader().setVisible(False)
        
        layout.addWidget(self.vehicles_table)
        tab.setLayout(layout)
        return tab
        
    def load_custom_vehicles(self):
        """Load custom vehicles into the table"""
        try:
            db = SessionLocal()
            vehicles = db.query(Vehicle).filter(Vehicle.is_custom == True).order_by(Vehicle.created_at.desc()).all()
            
            self.vehicles_table.setRowCount(len(vehicles))
            
            for row, vehicle in enumerate(vehicles):
                # Year, Make, Model
                self.vehicles_table.setItem(row, 0, QTableWidgetItem(vehicle.year))
                self.vehicles_table.setItem(row, 1, QTableWidgetItem(vehicle.make))
                self.vehicles_table.setItem(row, 2, QTableWidgetItem(vehicle.model))
                
                # Created By
                created_by = vehicle.created_by.first_name + " " + vehicle.created_by.last_name if vehicle.created_by else "Unknown"
                self.vehicles_table.setItem(row, 3, QTableWidgetItem(created_by))
                
                # Created At
                created_at = vehicle.created_at.strftime("%Y-%m-%d %H:%M") if vehicle.created_at else "Unknown"
                self.vehicles_table.setItem(row, 4, QTableWidgetItem(created_at))
                
                # Last Modified
                last_modified = vehicle.last_modified_at.strftime("%Y-%m-%d %H:%M") if vehicle.last_modified_at else "Never"
                self.vehicles_table.setItem(row, 5, QTableWidgetItem(last_modified))
                
                # Notes
                self.vehicles_table.setItem(row, 6, QTableWidgetItem(vehicle.notes or ""))
                
                # Actions
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(4, 4, 4, 4)
                actions_layout.setSpacing(8)
                
                edit_btn = QPushButton("Edit")
                edit_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #0078d4;
                        color: white;
                        border: none;
                        padding: 4px 12px;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #106ebe;
                    }
                """)
                edit_btn.clicked.connect(lambda checked, v=vehicle: self.edit_vehicle(v))
                
                delete_btn = QPushButton("Delete")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #d83b01;
                        color: white;
                        border: none;
                        padding: 4px 12px;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #ea4a1f;
                    }
                """)
                delete_btn.clicked.connect(lambda checked, v=vehicle: self.delete_vehicle(v))
                
                actions_layout.addWidget(edit_btn)
                actions_layout.addWidget(delete_btn)
                actions_layout.addStretch()
                
                self.vehicles_table.setCellWidget(row, 7, actions_widget)
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load custom vehicles: {str(e)}")
        finally:
            db.close()
            
    def edit_vehicle(self, vehicle):
        """Edit a custom vehicle"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Custom Vehicle")
        dialog.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
            }
            QLabel {
                color: white;
            }
            QLineEdit {
                background-color: #3d3d3d;
                color: white;
                padding: 8px;
                border: 1px solid #555555;
                border-radius: 4px;
            }
            QTextEdit {
                background-color: #3d3d3d;
                color: white;
                padding: 8px;
                border: 1px solid #555555;
                border-radius: 4px;
            }
        """)
        
        layout = QVBoxLayout(dialog)
        form = QFormLayout()
        
        # Create input fields
        year_input = QLineEdit(vehicle.year)
        make_input = QLineEdit(vehicle.make)
        model_input = QLineEdit(vehicle.model)
        notes_input = QTextEdit(vehicle.notes or "")
        
        form.addRow("Year:", year_input)
        form.addRow("Make:", make_input)
        form.addRow("Model:", model_input)
        form.addRow("Notes:", notes_input)
        
        layout.addLayout(form)
        
        # Add buttons
        button_layout = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #106ebe;
            }
        """)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #d83b01;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #ea4a1f;
            }
        """)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        save_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        if dialog.exec_() == QDialog.Accepted:
            try:
                db = SessionLocal()
                vehicle.year = year_input.text()
                vehicle.make = make_input.text()
                vehicle.model = model_input.text()
                vehicle.notes = notes_input.toPlainText()
                vehicle.last_modified_at = datetime.utcnow()
                vehicle.last_modified_by_id = self.current_user.id
                
                db.commit()
                self.load_custom_vehicles()  # Refresh the table
                QMessageBox.information(self, "Success", "Vehicle updated successfully!")
            except Exception as e:
                db.rollback()
                QMessageBox.critical(self, "Error", f"Failed to update vehicle: {str(e)}")
            finally:
                db.close()
                
    def delete_vehicle(self, vehicle):
        """Delete a custom vehicle"""
        try:
            db = SessionLocal()
            
            # Check for references in opportunities
            opps = db.query(Opportunity).filter(
                Opportunity.year == vehicle.year,
                Opportunity.make == vehicle.make,
                Opportunity.model == vehicle.model
            ).count()
            
            if opps > 0:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Cannot Delete Vehicle")
                msg.setInformativeText(f"This vehicle is referenced by {opps} opportunities and cannot be deleted.")
                msg.setWindowTitle("Cannot Delete")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.setDefaultButton(QMessageBox.Ok)
                msg.setStyleSheet("""
                    QMessageBox {
                        background-color: #2b2b2b;
                        color: white;
                    }
                    QMessageBox QLabel {
                        color: white;
                        font-size: 12px;
                        min-width: 300px;
                    }
                    QMessageBox QPushButton {
                        background-color: #0078d4;
                        color: white;
                        border: none;
                        padding: 6px 12px;
                        border-radius: 4px;
                        min-width: 80px;
                        font-weight: bold;
                    }
                    QMessageBox QPushButton:hover {
                        background-color: #106ebe;
                    }
                    QMessageBox QPushButton:pressed {
                        background-color: #005a9e;
                    }
                """)
                msg.exec_()
                return
            
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Are you sure you want to delete this vehicle?")
            msg.setInformativeText(f"Vehicle: {vehicle.year} {vehicle.make} {vehicle.model}")
            msg.setWindowTitle("Confirm Delete")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg.setDefaultButton(QMessageBox.No)
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: #2b2b2b;
                    color: white;
                }
                QMessageBox QLabel {
                    color: white;
                    font-size: 12px;
                    min-width: 300px;
                }
                QMessageBox QPushButton {
                    background-color: #0078d4;
                    color: white;
                    border: none;
                    padding: 6px 12px;
                    border-radius: 4px;
                    min-width: 80px;
                    font-weight: bold;
                }
                QMessageBox QPushButton:hover {
                    background-color: #106ebe;
                }
                QMessageBox QPushButton:pressed {
                    background-color: #005a9e;
                }
            """)
            
            if msg.exec_() == QMessageBox.Yes:
                db.delete(vehicle)
                db.commit()
                self.load_custom_vehicles()  # Refresh the table
                
                # Show success message
                success_msg = QMessageBox()
                success_msg.setIcon(QMessageBox.Information)
                success_msg.setText("Vehicle deleted successfully!")
                success_msg.setWindowTitle("Success")
                success_msg.setStandardButtons(QMessageBox.Ok)
                success_msg.setStyleSheet(msg.styleSheet())  # Reuse the same style
                success_msg.exec_()
                
        except Exception as e:
            db.rollback()
            error_msg = QMessageBox()
            error_msg.setIcon(QMessageBox.Critical)
            error_msg.setText("Error")
            error_msg.setInformativeText(f"Failed to delete vehicle: {str(e)}")
            error_msg.setWindowTitle("Error")
            error_msg.setStandardButtons(QMessageBox.Ok)
            error_msg.setStyleSheet(msg.styleSheet())  # Reuse the same style
            error_msg.exec_()
        finally:
            db.close()

    def export_to_excel(self):
        """Export ticket data to Excel with multiple sheets"""
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            
            # Get data from database
            db = SessionLocal()
            try:
                # Get all tickets
                all_tickets = db.query(Opportunity).all()
                
                # Separate tickets by status
                completed_tickets = [t for t in all_tickets if t.status.lower() == "completed"]
                in_progress_tickets = [t for t in all_tickets if t.status.lower() == "in progress"]
                needs_info_tickets = [t for t in all_tickets if t.status.lower() == "needs info"]
                new_tickets = [t for t in all_tickets if t.status.lower() == "new"]
                
                # Create sheets
                completed_sheet = wb.active
                completed_sheet.title = "Completed Tickets"
                in_progress_sheet = wb.create_sheet("In Progress Tickets")
                needs_info_sheet = wb.create_sheet("Needs Info Tickets")
                new_sheet = wb.create_sheet("New Tickets")
                
                # Define headers
                headers = [
                    "Ticket Number", "Year", "Make", "Model", 
                    "Creator", "Acceptor", "Date Created", 
                    "Date Completed", "Total Time", "Work Time"
                ]
                
                # Style for headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                # Function to write headers
                def write_headers(sheet):
                    for col, header in enumerate(headers, 1):
                        cell = sheet.cell(row=1, column=col)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                
                # Function to format duration
                def format_duration(duration):
                    if duration is None:
                        return "N/A"
                    total_seconds = int(duration.total_seconds())
                    days = total_seconds // 86400
                    hours = (total_seconds % 86400) // 3600
                    minutes = (total_seconds % 3600) // 60
                    return f"{days}d {hours}h {minutes}m"
                
                # Function to write ticket data
                def write_ticket_data(sheet, tickets):
                    write_headers(sheet)
                    for row, ticket in enumerate(tickets, 2):
                        sheet.cell(row=row, column=1).value = str(ticket.id)
                        sheet.cell(row=row, column=2).value = ticket.year
                        sheet.cell(row=row, column=3).value = ticket.make
                        sheet.cell(row=row, column=4).value = ticket.model
                        sheet.cell(row=row, column=5).value = f"{ticket.creator.first_name} {ticket.creator.last_name}"
                        sheet.cell(row=row, column=6).value = f"{ticket.acceptor.first_name} {ticket.acceptor.last_name}" if ticket.acceptor else "N/A"
                        sheet.cell(row=row, column=7).value = ticket.created_at.strftime("%Y-%m-%d %H:%M") if ticket.created_at else "N/A"
                        sheet.cell(row=row, column=8).value = ticket.completed_at.strftime("%Y-%m-%d %H:%M") if ticket.completed_at else "N/A"
                        sheet.cell(row=row, column=9).value = format_duration(ticket.response_time) if ticket.response_time else "N/A"
                        sheet.cell(row=row, column=10).value = format_duration(ticket.work_time) if ticket.work_time else "N/A"
                
                # Write data to each sheet
                write_ticket_data(completed_sheet, completed_tickets)
                write_ticket_data(in_progress_sheet, in_progress_tickets)
                write_ticket_data(needs_info_sheet, needs_info_tickets)
                write_ticket_data(new_sheet, new_tickets)
                
                # Auto-adjust column widths
                for sheet in wb.sheetnames:
                    worksheet = wb[sheet]
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Get save location from user
                file_name = f"SI_Opportunity_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Save Export File",
                    file_name,
                    "Excel Files (*.xlsx)"
                )
                
                if file_path:
                    wb.save(file_path)
                    QMessageBox.information(
                        self,
                        "Export Successful",
                        f"Data has been exported to:\n{file_path}"
                    )
                
            finally:
                db.close()
                
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"An error occurred while exporting data:\n{str(e)}"
            ) 